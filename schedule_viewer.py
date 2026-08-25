from __future__ import annotations

import html
import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles.colors import COLOR_INDEX


SCHEDULE_HEADERS = ("日期", "时间", "工作内容", "现场负责人", "人员安排")
SOURCE_HEADER_VARIANTS = {
    ("日期", "时间", "工作内容", "现场负责人", "人员安排"),
    ("日期", "时间", "生产内容", "现场负责人", "人员安排"),
}
NAME_EXCLUSIONS = {
    "无安排", "待安排", "负责人", "现场负责人", "人员安排", "工作内容", "全天", "白班", "夜班",
    "早班", "中班", "晚班", "休息", "培训", "确认", "复核", "其他", "生产", "设备",
    "冻干机", "外壁洗", "洗烘", "灌装", "包装", "提前出", "待确认",
}
COMMON_SURNAMES = set(
    "赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜戚谢邹喻柏水窦章云苏潘葛奚范彭郎鲁韦昌马苗凤花方俞任袁柳鲍史唐费廉岑薛雷贺倪汤滕殷罗毕郝邬安常乐于时傅皮卞齐康伍余元顾孟平黄和穆萧尹姚邵湛汪祁毛禹狄米贝明臧计伏成戴谈宋茅庞熊纪舒屈项祝董梁杜阮蓝闵席季麻强贾路娄危江童颜郭梅盛林刁钟徐邱骆高夏蔡田樊胡凌霍虞万支柯管卢莫房裘缪干解应宗丁宣贲邓郁单杭洪包诸左石崔吉龚程嵇邢裴陆荣翁荀羊惠甄曲封芮羿储靳汲邴糜松井段富巫乌焦巴弓牧隗山谷车侯宓蓬全郗班仰秋仲伊宫宁仇栾暴甘厉戎祖武符刘景詹束龙叶幸司韶郜黎蓟薄印宿白怀蒲邰从鄂索咸籍赖卓蔺屠蒙池乔阴胥能苍双闻莘党翟谭贡劳逄姬申扶堵冉宰郦雍却璩桑桂濮牛寿通边扈燕冀郏浦尚农温别庄晏柴瞿阎充慕连茹习宦艾鱼容向古易慎戈廖庾终暨居衡步都耿满弘匡国文寇广禄阙东欧殳沃利蔚越夔隆师巩厍聂晁勾敖融冷訾辛阚那简饶空曾毋沙乜养鞠须丰巢关蒯相查后荆红游竺权逯盖益桓公"
)
COMMON_SURNAMES.update("覃")


@dataclass
class ScheduleWorkbook:
    workbook: object
    schedule_sheets: list[str]


def load_schedule_workbook(file_bytes: bytes) -> ScheduleWorkbook:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheets = [sheet.title for sheet in workbook.worksheets if _find_header_row(sheet) is not None]
    if not sheets:
        raise ValueError("未找到包含“日期、时间、工作内容、现场负责人、人员安排”的排班工作表。")
    return ScheduleWorkbook(workbook=workbook, schedule_sheets=sheets)


def _find_header_row(sheet):
    for row in range(1, min(sheet.max_row, 12) + 1):
        values = [str(sheet.cell(row, col).value or "").strip() for col in range(1, min(sheet.max_column, 12) + 1)]
        for start in range(0, max(1, len(values) - len(SCHEDULE_HEADERS) + 1)):
            if tuple(values[start : start + len(SCHEDULE_HEADERS)]) in SOURCE_HEADER_VARIANTS:
                return row, start + 1
    return None


def sheet_layout(sheet):
    location = _find_header_row(sheet)
    if location is None:
        raise ValueError("所选工作表不是可识别的生产排班表。")
    header_row, first_col = location
    last_col = first_col + len(SCHEDULE_HEADERS) - 1
    last_row = max(
        row
        for row in range(header_row, sheet.max_row + 1)
        if any(sheet.cell(row, col).value not in (None, "") for col in range(first_col, last_col + 1))
    )
    return header_row, first_col, last_col, last_row


def _display_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value)


def _split_joined_chinese_names(token):
    """Split a punctuation-free Chinese name run using surname boundaries."""
    if not re.fullmatch(r"[\u4e00-\u9fff]+", token) or len(token) <= 3:
        return [token]
    if len(token) == 4 and token[0] in COMMON_SURNAMES and token[2] in COMMON_SURNAMES:
        return [token[:2], token[2:]]
    if len(token) == 4:
        return [token]
    size = len(token)
    best = [None] * (size + 1)
    best[0] = []
    for start in range(size):
        if best[start] is None or token[start] not in COMMON_SURNAMES:
            continue
        for length in (3, 2, 4):
            end = start + length
            if end <= size and (end == size or token[end] in COMMON_SURNAMES):
                candidate = best[start] + [token[start:end]]
                if best[end] is None or abs(length - 3) + len(candidate) < len(best[end]) + 1:
                    best[end] = candidate
    return best[size] or [token]


def normalize_people_text(value):
    """Normalize separators and split clearly concatenated Chinese names."""
    text = _display_value(value).strip()
    if not text:
        return ""
    parts = re.split(r"[、，,；;\/\n\t]+", text)
    normalized = []
    for part in parts:
        part = re.sub(r"\s+", " ", part).strip(" 、，,；;")
        if not part:
            continue
        if re.fullmatch(r"[\u4e00-\u9fff]+", part):
            normalized.extend(_split_joined_chinese_names(part))
        else:
            normalized.append(part)
    return "、".join(dict.fromkeys(normalized))


def _merged_value(sheet, row, col):
    cell = sheet.cell(row, col)
    if not isinstance(cell, MergedCell):
        return _display_value(cell.value)
    for merged in sheet.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return _display_value(sheet.cell(merged.min_row, merged.min_col).value)
    return ""


def schedule_records(sheet):
    header_row, first_col, _, last_row = sheet_layout(sheet)
    records = []
    for row in range(header_row + 1, last_row + 1):
        values = [_merged_value(sheet, row, first_col + offset).strip() for offset in range(5)]
        values[3] = normalize_people_text(values[3])
        values[4] = normalize_people_text(values[4])
        if any(values):
            records.append({"row": row, **dict(zip(SCHEDULE_HEADERS, values))})
    return records


def available_people(sheet):
    candidates = set()
    for record in schedule_records(sheet):
        for field in ("现场负责人", "人员安排"):
            text = record[field]
            for token in re.split(r"[、，,；;\n/（）()：:\s]+", text):
                token = token.strip()
                if (
                    re.fullmatch(r"[\u4e00-\u9fff]{2,4}", token)
                    and token not in NAME_EXCLUSIONS
                    and not any(word in token for word in ("人员", "安排", "现场", "负责", "清洗", "灌装", "冻干", "设备", "提前"))
                ):
                    candidates.add(token)
    return sorted(candidates)


def matching_records(sheet, person):
    if not person:
        return []
    return [
        record
        for record in schedule_records(sheet)
        if person in record["现场负责人"] or person in record["人员安排"]
    ]


def generated_schedule_records(result):
    """Convert the scheduler result into the same five-column read-only view."""
    if result is None or result.empty:
        return []
    assigned = result[result["status"] == "已排班"].copy()
    records = []
    group_columns = ["date", "production_line", "shift", "batch_id", "task_type", "required_role"]
    for values, group in assigned.groupby(group_columns, dropna=False, sort=True):
        schedule_date, line, shift, batch, task, role = values
        names = "、".join(dict.fromkeys(str(name) for name in group["employee_name"] if str(name).strip()))
        role_text = "" if role is None else str(role)
        records.append({
            "row": len(records) + 2,
            "日期": schedule_date.strftime("%Y-%m-%d") if hasattr(schedule_date, "strftime") else str(schedule_date),
            "时间": str(shift),
            "工作内容": " / ".join(part for part in (str(line), str(batch), str(task)) if part and part != "nan"),
            "现场负责人": names if "负责人" in role_text else "",
            "人员安排": "" if "负责人" in role_text else names,
        })
    return records


def people_from_records(records):
    candidates = set()
    for record in records:
        for field in ("现场负责人", "人员安排"):
            for token in re.split(r"[、，,；;\n/（）()：:\s]+", record.get(field, "")):
                if re.fullmatch(r"[\u4e00-\u9fff]{2,4}", token or ""):
                    candidates.add(token)
    return sorted(candidates)


def render_records_html(records, selected_person=""):
    header = "".join(f'<td class="schedule-cell schedule-header">{label}</td>' for label in SCHEDULE_HEADERS)
    rows = [f"<tr>{header}</tr>"]
    for record in records:
        match_type = _record_match_type(record, selected_person)
        cells = []
        for index, label in enumerate(SCHEDULE_HEADERS):
            classes = "schedule-cell"
            if match_type == "participant":
                classes += " schedule-match"
            elif match_type == "leader" and index == 0:
                classes += " schedule-leader-match"
            cells.append(f'<td class="{classes}">{html.escape(str(record.get(label, ""))).replace(chr(10), "<br>")}</td>')
        rows.append("<tr>" + "".join(cells) + "</tr>")
    return f"""
    <div class="schedule-wrap"><table class="schedule-table">
      <colgroup><col class="date-col"><col class="time-col"><col class="work-col"><col class="lead-col"><col class="people-col"></colgroup>
      {''.join(rows)}
    </table></div>
    """


def _pdf_fonts():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.pdfbase.ttfonts import TTFont

    registered = set(pdfmetrics.getRegisteredFontNames())
    chinese_name = "MicrosoftYaHei"
    english_name = "Arial"
    if chinese_name not in registered:
        chinese_paths = [
            Path("C:/Windows/Fonts/msyh.ttc"),
            Path("/usr/share/fonts/truetype/msttcorefonts/msyh.ttf"),
        ]
        for path in chinese_paths:
            if path.exists():
                try:
                    pdfmetrics.registerFont(TTFont(chinese_name, str(path)))
                    break
                except Exception:
                    continue
        else:
            chinese_name = "STSong-Light"
            if chinese_name not in registered:
                pdfmetrics.registerFont(UnicodeCIDFont(chinese_name))
    if english_name not in registered:
        arial_paths = [Path("C:/Windows/Fonts/arial.ttf"), Path("/usr/share/fonts/truetype/msttcorefonts/Arial.ttf")]
        for path in arial_paths:
            if path.exists():
                try:
                    pdfmetrics.registerFont(TTFont(english_name, str(path)))
                    break
                except Exception:
                    continue
        else:
            english_name = "Helvetica"
    return chinese_name, english_name


def _pdf_markup(value, chinese_font, english_font):
    rendered_lines = []
    for line in str(value or "").splitlines() or [""]:
        escaped = html.escape(line)
        rendered_lines.append(re.sub(
            r"([A-Za-z0-9][A-Za-z0-9 ._+:/()%-]*)",
            lambda match: f'<font name="{english_font}">{match.group(1)}</font>',
            escaped,
        ))
    return "<br/>".join(rendered_lines)


def _pdf_layout(records, available_width):
    row_count = max(1, len(records))
    if row_count <= 24:
        font_size = 8.5
    elif row_count <= 40:
        font_size = 7.2
    elif row_count <= 60:
        font_size = 5.8
    elif row_count <= 80:
        font_size = 4.8
    else:
        font_size = 4.2
    leading = font_size * 1.18
    padding = max(0.8, min(2.8, 56 / row_count))

    def visual_length(value):
        return sum(2 if "\u4e00" <= char <= "\u9fff" else 1 for char in str(value or ""))

    maxima = []
    for label in SCHEDULE_HEADERS:
        lengths = [visual_length(label)] + [visual_length(record.get(label, "")) for record in records]
        maxima.append(min(max(lengths), 60))
    minimums = [52, 54, 142, 62, 92]
    remaining = max(0, available_width - sum(minimums))
    pressures = [maxima[0] * 0.5, maxima[1] * 0.7, maxima[2] * 1.8, maxima[3], maxima[4] * 1.4]
    pressure_total = sum(pressures) or 1
    widths = [minimum + remaining * pressure / pressure_total for minimum, pressure in zip(minimums, pressures)]
    return widths, font_size, leading, padding


def export_records_pdf(records, title, selected_person=""):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfgen import canvas

    chinese_font, english_font = _pdf_fonts()
    buffer = io.BytesIO()
    page_width, page_height = A4
    margin_x = 7 * mm
    margin_y = 7 * mm
    table_width = page_width - margin_x * 2
    title_height = 17
    note_height = 10 if selected_person else 0
    gap_height = 6
    available_table_height = page_height - margin_y * 2 - title_height - note_height - gap_height
    widths, initial_size, _, _ = _pdf_layout(records, table_width)

    def font_for_char(char):
        return chinese_font if "\u4e00" <= char <= "\u9fff" else english_font

    def text_width(text, size):
        return sum(pdfmetrics.stringWidth(char, font_for_char(char), size) for char in text)

    def wrap_text(value, width, size):
        lines = []
        for source_line in str(value or "").splitlines() or [""]:
            current = ""
            current_width = 0.0
            for char in source_line:
                char_width = pdfmetrics.stringWidth(char, font_for_char(char), size)
                if current and current_width + char_width > width:
                    lines.append(current)
                    current = char
                    current_width = char_width
                else:
                    current += char
                    current_width += char_width
            lines.append(current)
        return lines or [""]

    def row_layouts(size):
        leading = size * 1.15
        padding = max(0.6, size * 0.22)
        layouts = []
        header_lines = [[label] for label in SCHEDULE_HEADERS]
        layouts.append((header_lines, leading + padding * 2 + 1))
        for record in records:
            cell_lines = [wrap_text(record.get(label, ""), widths[index] - 4, size) for index, label in enumerate(SCHEDULE_HEADERS)]
            height = max(len(lines) for lines in cell_lines) * leading + padding * 2
            layouts.append((cell_lines, height))
        return layouts, leading, padding

    font_size = initial_size
    layouts, leading, padding = row_layouts(font_size)
    while sum(height for _, height in layouts) > available_table_height and font_size > 3.0:
        font_size = round(font_size - 0.2, 2)
        layouts, leading, padding = row_layouts(font_size)
    total_height = sum(height for _, height in layouts)
    if total_height > available_table_height:
        ratio = available_table_height / total_height
        layouts = [(lines, height * ratio) for lines, height in layouts]
        leading *= ratio

    pdf = canvas.Canvas(buffer, pagesize=A4, pageCompression=1)
    pdf.setTitle(title)

    def draw_mixed_line(text, x, baseline, size, align="left", max_width=None):
        width = text_width(text, size)
        if align == "center":
            cursor = x - width / 2
        elif align == "right":
            cursor = x - width
        else:
            cursor = x
        if max_width is not None and width > max_width:
            return
        for char in text:
            font = font_for_char(char)
            pdf.setFont(font, size)
            pdf.drawString(cursor, baseline, char)
            cursor += pdfmetrics.stringWidth(char, font, size)

    pdf.setFillColor(colors.HexColor("#17365D"))
    draw_mixed_line(title, page_width / 2, page_height - margin_y - 11, 12, align="center")
    table_top = page_height - margin_y - title_height
    if selected_person:
        pdf.setFillColor(colors.HexColor("#3F4B5B"))
        note = f"人员定位：{selected_person}（黄色为参与班次，蓝色日期为现场负责人）"
        draw_mixed_line(note, margin_x, table_top - 7, 5.5)
        table_top -= note_height

    y = table_top
    all_rows = [{label: label for label in SCHEDULE_HEADERS}] + records
    for row_index, (record, (cell_lines, row_height)) in enumerate(zip(all_rows, layouts)):
        y -= row_height
        match_type = "" if row_index == 0 else _record_match_type(record, selected_person)
        if row_index == 0:
            pdf.setFillColor(colors.HexColor("#D9EAF7"))
            pdf.rect(margin_x, y, table_width, row_height, fill=1, stroke=0)
        elif match_type == "participant":
            pdf.setFillColor(colors.HexColor("#FFF2B2"))
            pdf.rect(margin_x, y, table_width, row_height, fill=1, stroke=0)
        elif match_type == "leader":
            pdf.setFillColor(colors.HexColor("#DCEBFF"))
            pdf.rect(margin_x, y, widths[0], row_height, fill=1, stroke=0)

        x = margin_x
        for column_index, (width, lines) in enumerate(zip(widths, cell_lines)):
            pdf.saveState()
            clip = pdf.beginPath()
            clip.rect(x + 0.5, y + 0.5, width - 1, row_height - 1)
            pdf.clipPath(clip, stroke=0, fill=0)
            text_block_height = len(lines) * leading
            baseline = y + (row_height + text_block_height) / 2 - leading + (leading - font_size) / 2
            pdf.setFillColor(colors.HexColor("#17365D") if row_index == 0 else colors.HexColor("#263142"))
            for line in lines:
                if column_index == 2:
                    draw_mixed_line(line, x + 2, baseline, font_size, align="left", max_width=width - 4)
                else:
                    draw_mixed_line(line, x + width / 2, baseline, font_size, align="center", max_width=width - 4)
                baseline -= leading
            pdf.restoreState()
            x += width

        pdf.setStrokeColor(colors.HexColor("#AEB8C6"))
        pdf.setLineWidth(0.25)
        pdf.line(margin_x, y, margin_x + table_width, y)

    pdf.setStrokeColor(colors.HexColor("#AEB8C6"))
    pdf.setLineWidth(0.3)
    x = margin_x
    pdf.line(margin_x, table_top, margin_x + table_width, table_top)
    for width in [0] + widths:
        pdf.line(x, y, x, table_top)
        x += width
    pdf.line(margin_x + table_width, y, margin_x + table_width, table_top)
    pdf.showPage()
    pdf.save()
    return buffer.getvalue()


def _record_match_type(record, person):
    if not person:
        return ""
    is_leader = person in record["现场负责人"]
    is_participant = person in record["人员安排"]
    if is_participant:
        return "participant"
    if is_leader:
        return "leader"
    return ""


def _color_rgb(color, fallback):
    if color is None:
        return fallback
    if color.type == "rgb" and color.rgb:
        return f"#{color.rgb[-6:]}"
    if color.type == "indexed" and color.indexed is not None and color.indexed < len(COLOR_INDEX):
        return f"#{COLOR_INDEX[color.indexed][-6:]}"
    return fallback


def render_schedule_html(sheet, selected_person=""):
    # The viewer uses normalized records instead of source cell formatting so
    # every uploaded workbook gets consistent spacing, alignment and wrapping.
    return render_records_html(schedule_records(sheet), selected_person)

    # Legacy source-style renderer retained below for compatibility reference.
    header_row, first_col, last_col, last_row = sheet_layout(sheet)
    records_by_row = {record["row"]: record for record in schedule_records(sheet)}
    # A selected-person view deliberately expands merged cells. This prevents a
    # date cell shared by several tasks from making unrelated tasks look selected.
    if selected_person:
        rows = []
        header = "".join(f'<td class="schedule-cell schedule-header">{label}</td>' for label in SCHEDULE_HEADERS)
        rows.append(f"<tr>{header}</tr>")
        for record in records_by_row.values():
            match_type = _record_match_type(record, selected_person)
            cells = []
            for index, label in enumerate(SCHEDULE_HEADERS):
                classes = "schedule-cell"
                if match_type == "participant":
                    classes += " schedule-match"
                elif match_type == "leader" and index == 0:
                    classes += " schedule-leader-match"
                cells.append(f'<td class="{classes}">{html.escape(record[label]).replace(chr(10), "<br>")}</td>')
            rows.append("<tr>" + "".join(cells) + "</tr>")
        return f"""
        <div class="schedule-wrap">
          <table class="schedule-table">
            <colgroup><col class="date-col"><col class="time-col"><col class="work-col"><col class="lead-col"><col class="people-col"></colgroup>
            {''.join(rows)}
          </table>
        </div>
        """

    merged_starts = {}
    merged_covered = set()
    for merged in sheet.merged_cells.ranges:
        if merged.max_row < header_row or merged.min_row > last_row or merged.max_col < first_col or merged.min_col > last_col:
            continue
        start = (max(merged.min_row, header_row), max(merged.min_col, first_col))
        end_row = min(merged.max_row, last_row)
        end_col = min(merged.max_col, last_col)
        merged_starts[start] = (end_row - start[0] + 1, end_col - start[1] + 1)
        for row in range(start[0], end_row + 1):
            for col in range(start[1], end_col + 1):
                if (row, col) != start:
                    merged_covered.add((row, col))

    rows = []
    for row in range(header_row, last_row + 1):
        cells = []
        for col in range(first_col, last_col + 1):
            if (row, col) in merged_covered:
                continue
            source = sheet.cell(row, col)
            value = _display_value(source.value)
            rowspan, colspan = merged_starts.get((row, col), (1, 1))
            fill = _color_rgb(source.fill.fgColor, "#ffffff") if source.fill.fill_type else "#ffffff"
            color = _color_rgb(source.font.color, "#263142") if source.font.color else "#263142"
            align = source.alignment.horizontal or ("center" if col != first_col + 2 else "left")
            weight = "700" if source.font.bold or row == header_row else "400"
            classes = "schedule-cell"
            if row == header_row:
                classes += " schedule-header"
            attrs = f' rowspan="{rowspan}" colspan="{colspan}"' if rowspan > 1 or colspan > 1 else ""
            style = f"background:{fill};color:{color};text-align:{align};font-weight:{weight};"
            cells.append(
                f'<td class="{classes}"{attrs} style="{style}">{html.escape(value).replace(chr(10), "<br>")}</td>'
            )
        rows.append("<tr>" + "".join(cells) + "</tr>")

    return f"""
    <div class="schedule-wrap">
      <table class="schedule-table">
        <colgroup><col class="date-col"><col class="time-col"><col class="work-col"><col class="lead-col"><col class="people-col"></colgroup>
        {''.join(rows)}
      </table>
    </div>
    """


def _pdf_fonts_v2():
    """Register clear sans-serif fonts on Windows and Streamlit/Linux."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.pdfbase.ttfonts import TTFont

    def first_available(name, paths):
        if name in pdfmetrics.getRegisteredFontNames():
            return name
        for value in paths:
            path = Path(value)
            if not path.exists():
                continue
            try:
                pdfmetrics.registerFont(TTFont(name, str(path)))
                return name
            except Exception:
                continue
        return ""

    cn = first_available("CMFSansCN", [
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simhei.ttf",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJKsc-Regular.otf",
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
    ])
    cn_bold = first_available("CMFSansCNBold", [
        "C:/Windows/Fonts/msyhbd.ttc",
        "C:/Windows/Fonts/simhei.ttf",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJKsc-Bold.otf",
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
    ])
    if not cn:
        cn = "HeiseiKakuGo-W5"
        if cn not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(UnicodeCIDFont(cn))
    cn_bold = cn_bold or cn

    en = first_available("CMFSansEN", [
        "C:/Windows/Fonts/arial.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]) or "Helvetica"
    en_bold = first_available("CMFSansENBold", [
        "C:/Windows/Fonts/arialbd.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    ]) or "Helvetica-Bold"
    return cn, cn_bold, en, en_bold


def _export_records_pdf_v2(records, title, selected_person=""):
    """Render an A4 single-page schedule in the source workbook's compact style."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfgen import canvas

    cn, cn_bold, en, en_bold = _pdf_fonts_v2()
    page_width, page_height = A4
    margin_x, margin_y = 17 * mm, 14 * mm
    table_width = page_width - 2 * margin_x
    widths = [table_width * ratio for ratio in (0.13, 0.13, 0.39, 0.13, 0.22)]
    title_space = 25
    note_space = 12 if selected_person else 0
    table_top = page_height - margin_y - title_space - note_space
    table_capacity = table_top - margin_y
    buffer = io.BytesIO()

    def face(char, bold=False):
        codepoint = ord(char)
        is_cjk = (
            0x3000 <= codepoint <= 0x303F  # CJK punctuation
            or 0x3400 <= codepoint <= 0x9FFF  # CJK ideographs
            or 0xFF00 <= codepoint <= 0xFFEF  # full-width forms
        )
        return (cn_bold if bold else cn) if is_cjk else (en_bold if bold else en)

    def measure(text, size, bold=False):
        return sum(pdfmetrics.stringWidth(char, face(char, bold), size) for char in str(text))

    def wrap(value, width, size):
        result = []
        for source in str(value or "").splitlines() or [""]:
            current, used = "", 0.0
            for char in source:
                char_width = pdfmetrics.stringWidth(char, face(char), size)
                if current and used + char_width > width:
                    result.append(current)
                    current, used = char, char_width
                else:
                    current += char
                    used += char_width
            result.append(current)
        return result or [""]

    def display_values(index, record):
        values = [str(record.get(label, "") or "") for label in SCHEDULE_HEADERS]
        if index == 0:
            return values
        previous = records[index - 1]
        same_date = values[0] == str(previous.get(SCHEDULE_HEADERS[0], "") or "")
        if same_date:
            values[0] = ""
            if values[1] == str(previous.get(SCHEDULE_HEADERS[1], "") or ""):
                values[1] = ""
            if values[3] == str(previous.get(SCHEDULE_HEADERS[3], "") or ""):
                values[3] = ""
        return values

    def build_layout(size):
        leading = size * 1.18
        padding = max(0.65, size * 0.23)
        header_height = leading + 2 * padding + 2
        rows = []
        for index, record in enumerate(records):
            values = display_values(index, record)
            cells = [wrap(value, widths[col] - 6, size) for col, value in enumerate(values)]
            height = max(len(lines) for lines in cells) * leading + 2 * padding
            rows.append((cells, height))
        return leading, padding, header_height, rows

    size = 7.7 if len(records) <= 42 else 6.8 if len(records) <= 65 else 5.9
    leading, padding, header_height, row_layouts = build_layout(size)
    while header_height + sum(height for _, height in row_layouts) > table_capacity and size > 4.3:
        size = round(size - 0.2, 2)
        leading, padding, header_height, row_layouts = build_layout(size)
    total_height = header_height + sum(height for _, height in row_layouts)
    if total_height > table_capacity:
        ratio = table_capacity / total_height
        header_height *= ratio
        row_layouts = [(cells, height * ratio) for cells, height in row_layouts]
        leading *= ratio

    pdf = canvas.Canvas(buffer, pagesize=A4, pageCompression=1)
    pdf.setTitle(str(title))

    def draw_line(text, x, baseline, font_size, align="left", bold=False):
        text = str(text)
        line_width = measure(text, font_size, bold)
        cursor = x - line_width / 2 if align == "center" else x - line_width if align == "right" else x
        for char in text:
            font = face(char, bold)
            pdf.setFont(font, font_size)
            pdf.drawString(cursor, baseline, char)
            cursor += pdfmetrics.stringWidth(char, font, font_size)

    pdf.setFillColor(colors.HexColor("#111827"))
    draw_line(title, page_width / 2, page_height - margin_y - 14, 12.5, "center", True)
    if selected_person:
        pdf.setFillColor(colors.HexColor("#4B5563"))
        draw_line(f"人员定位：{selected_person}（浅黄：参与班次；浅蓝日期：现场负责人）", margin_x, table_top + 4, 6.2)

    y = table_top - header_height
    pdf.setFillColor(colors.HexColor("#F3F4F6"))
    pdf.rect(margin_x, y, table_width, header_height, fill=1, stroke=0)
    x = margin_x
    for label, width in zip(SCHEDULE_HEADERS, widths):
        pdf.setFillColor(colors.HexColor("#111827"))
        draw_line(label, x + width / 2, y + (header_height - size) / 2, size, "center", True)
        x += width
    pdf.setStrokeColor(colors.HexColor("#1F2937"))
    pdf.setLineWidth(0.6)
    pdf.line(margin_x, table_top, margin_x + table_width, table_top)
    pdf.line(margin_x, y, margin_x + table_width, y)

    for index, (record, (cells, row_height)) in enumerate(zip(records, row_layouts)):
        row_top = y
        y -= row_height
        match_type = _record_match_type(record, selected_person)
        if match_type == "participant":
            pdf.setFillColor(colors.HexColor("#FFF4BF"))
            pdf.rect(margin_x, y, table_width, row_height, fill=1, stroke=0)
        elif match_type == "leader":
            pdf.setFillColor(colors.HexColor("#DCEEFF"))
            pdf.rect(margin_x, y, widths[0], row_height, fill=1, stroke=0)

        x = margin_x
        for column, (cell_lines, width) in enumerate(zip(cells, widths)):
            text_height = len(cell_lines) * leading
            baseline = y + (row_height + text_height) / 2 - leading + (leading - size) / 2
            pdf.setFillColor(colors.HexColor("#111827"))
            for line in cell_lines:
                if column == 2:
                    draw_line(line, x + 3, baseline, size)
                else:
                    draw_line(line, x + width / 2, baseline, size, "center")
                baseline -= leading
            x += width

        previous = records[index - 1] if index else None
        date_changed = previous is None or record.get(SCHEDULE_HEADERS[0], "") != previous.get(SCHEDULE_HEADERS[0], "")
        time_changed = date_changed or record.get(SCHEDULE_HEADERS[1], "") != previous.get(SCHEDULE_HEADERS[1], "")
        line_start = margin_x if date_changed else margin_x + widths[0] if time_changed else margin_x + widths[0] + widths[1]
        pdf.setStrokeColor(colors.HexColor("#6B7280"))
        pdf.setLineWidth(0.45 if date_changed else 0.22)
        pdf.line(line_start, y, margin_x + table_width, y)

    pdf.setStrokeColor(colors.HexColor("#1F2937"))
    pdf.setLineWidth(0.6)
    pdf.line(margin_x, y, margin_x + table_width, y)
    pdf.showPage()
    pdf.save()
    return buffer.getvalue()


# Use the refined renderer without changing callers or schedule data structures.
export_records_pdf = _export_records_pdf_v2


def export_schedule_pdf(sheet, selected_person=""):
    title = str(sheet.cell(2, 2).value or f"{sheet.title} 生产排班")
    return export_records_pdf(schedule_records(sheet), title, selected_person)
